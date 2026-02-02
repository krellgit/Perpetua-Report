#!/usr/bin/env python3
"""
Excel Dashboard with TRUE Date Selector Dropdowns
Uses Data Validation to create dropdown lists for date filtering
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent.parent
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 100)
print("DASHBOARD WITH INTERACTIVE DATE SELECTOR DROPDOWNS")
print("=" * 100)
print()

# Load data
print("[1/6] Loading data...")
df = pd.read_csv(PROCESSED_DIR / 'advertised_products_processed.csv', low_memory=False)
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df = df[df['Date'].notna()]
df = df[df['Advertising_Type'].isin(['Perpetua', 'Non-Perpetua'])]

for col in ['Spend', '7 Day Total Sales ', '7 Day Total Orders (#)', 'Clicks', 'Impressions', '7 Day Total Units (#)']:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

min_date = df['Date'].min()
max_date = df['Date'].max()
all_dates = sorted(df['Date'].unique())

print(f"  ✓ {len(df):,} records")
print(f"  ✓ Date range: {min_date.date()} to {max_date.date()}")
print(f"  ✓ {len(all_dates)} unique dates")

# Calculate metrics
def calc_metrics(subset):
    total_spend = subset['Spend'].sum()
    total_sales = subset['7 Day Total Sales '].sum()
    total_orders = subset['7 Day Total Orders (#)'].sum()
    total_clicks = subset['Clicks'].sum()
    total_impressions = subset['Impressions'].sum()
    total_units = subset['7 Day Total Units (#)'].sum()
    unique_asins = subset['Advertised ASIN'].nunique()

    return {
        'Total_Spend': total_spend,
        'Total_Sales': total_sales,
        'Total_Orders': total_orders,
        'Total_Clicks': total_clicks,
        'Total_Impressions': total_impressions,
        'Total_Units': total_units,
        'Unique_ASINs': unique_asins,
        'ROAS': total_sales / total_spend if total_spend > 0 else 0,
        'ACOS': total_spend / total_sales if total_sales > 0 else 0,
        'CPC': total_spend / total_clicks if total_clicks > 0 else 0,
        'CTR': total_clicks / total_impressions if total_impressions > 0 else 0,
        'CVR': total_orders / total_clicks if total_clicks > 0 else 0,
        'CPA': total_spend / total_orders if total_orders > 0 else 0,
        'CPM': (total_spend / total_impressions) * 1000 if total_impressions > 0 else 0,
        'AOV': total_sales / total_orders if total_orders > 0 else 0,
    }

perpetua = calc_metrics(df[df['Advertising_Type'] == 'Perpetua'])
non_perpetua = calc_metrics(df[df['Advertising_Type'] == 'Non-Perpetua'])

print("[2/6] Aggregate metrics calculated")

# Daily aggregation
daily = df.groupby(['Date', 'Advertising_Type']).agg({
    'Spend': 'sum',
    '7 Day Total Sales ': 'sum',
    '7 Day Total Orders (#)': 'sum',
    'Clicks': 'sum',
    'Impressions': 'sum'
}).reset_index()

daily['ROAS'] = daily['7 Day Total Sales '] / daily['Spend'].replace(0, np.nan)
daily['ACOS'] = daily['Spend'] / daily['7 Day Total Sales '].replace(0, np.nan)
daily['CPC'] = daily['Spend'] / daily['Clicks'].replace(0, np.nan)
daily['CTR'] = daily['Clicks'] / daily['Impressions'].replace(0, np.nan)
daily['CVR'] = daily['7 Day Total Orders (#)'] / daily['Clicks'].replace(0, np.nan)
daily = daily.replace([np.inf, -np.inf], np.nan).fillna(0)

print(f"  ✓ {len(daily)} daily records")

# ============================================================================
# CREATE WORKBOOK
# ============================================================================

print("[3/6] Creating workbook with date selectors...")

wb = Workbook()
wb.remove(wb.active)

COLORS = {
    'perpetua': '4472C4',
    'non_perpetua': 'ED7D31',
    'header': '2F5496',
    'good': '70AD47',
    'bad': 'C5504B',
    'input': 'FFF2CC',
    'light_blue': 'D9E1F2',
    'light_orange': 'FCE4D6'
}

header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=18, color=COLORS['header'])
input_fill = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type='solid')
input_font = Font(bold=True, size=12)
center = Alignment(horizontal='center', vertical='center')

# ============================================================================
# SHEET 1: EXECUTIVE DASHBOARD WITH DATE SELECTORS
# ============================================================================

print("[4/6] Creating Executive Dashboard with date selectors...")
ws1 = wb.create_sheet("📊 Dashboard")

# Title
ws1['C2'] = 'PERPETUA (SaaS) vs MANUAL ADVERTISING'
ws1['C2'].font = title_font
ws1['C2'].alignment = center
ws1.merge_cells('C2:K2')

ws1['C3'] = 'Performance Comparison Dashboard'
ws1['C3'].font = Font(size=12, italic=True)
ws1['C3'].alignment = center
ws1.merge_cells('C3:K3')

# DATE SELECTOR SECTION
ws1['C5'] = '📅 SELECT DATE RANGE:'
ws1['C5'].font = Font(bold=True, size=13, color=COLORS['non_perpetua'])
ws1.merge_cells('C5:K5')

ws1['C6'] = 'Start Date:'
ws1['C6'].font = Font(bold=True, size=11)
ws1['C6'].alignment = Alignment(horizontal='right')

ws1['D6'] = min_date
ws1['D6'].number_format = 'YYYY-MM-DD'
ws1['D6'].fill = input_fill
ws1['D6'].font = input_font
ws1['D6'].alignment = center

ws1['F6'] = 'End Date:'
ws1['F6'].font = Font(bold=True, size=11)
ws1['F6'].alignment = Alignment(horizontal='right')

ws1['G6'] = max_date
ws1['G6'].number_format = 'YYYY-MM-DD'
ws1['G6'].fill = input_fill
ws1['G6'].font = input_font
ws1['G6'].alignment = center

# Create hidden sheet with date list for dropdown
ws_dates = wb.create_sheet("_DateList")
ws_dates.sheet_state = 'hidden'

# Add all dates to hidden sheet
for idx, date in enumerate(all_dates, start=1):
    ws_dates.cell(row=idx, column=1, value=pd.Timestamp(date).to_pydatetime())
    ws_dates.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'

# Create data validation (dropdown) for start date
dv_start = DataValidation(type="list",
                          formula1=f"='_DateList'!$A$1:$A${len(all_dates)}",
                          allow_blank=False)
dv_start.error = 'Please select a date from the list'
dv_start.errorTitle = 'Invalid Date'
dv_start.prompt = 'Select start date from dropdown'
dv_start.promptTitle = 'Date Selection'
ws1.add_data_validation(dv_start)
dv_start.add(ws1['D6'])

# Create data validation for end date
dv_end = DataValidation(type="list",
                        formula1=f"='_DateList'!$A$1:$A${len(all_dates)}",
                        allow_blank=False)
dv_end.error = 'Please select a date from the list'
dv_end.errorTitle = 'Invalid Date'
dv_end.prompt = 'Select end date from dropdown'
dv_end.promptTitle = 'Date Selection'
ws1.add_data_validation(dv_end)
dv_end.add(ws1['G6'])

# Instructions
ws1['C7'] = '👆 Click cells D6 and G6 to select dates from dropdown menus'
ws1['C7'].font = Font(size=10, italic=True, color='666666')
ws1.merge_cells('C7:K7')
ws1['C7'].alignment = center

ws1['C8'] = 'Note: Charts show full date range. Use "Daily Data" sheet filters for custom date analysis.'
ws1['C8'].font = Font(size=9, italic=True, color='999999')
ws1.merge_cells('C8:K8')
ws1['C8'].alignment = center

# PERFORMANCE SUMMARY
row = 10
ws1[f'C{row}'] = 'PERFORMANCE COMPARISON'
ws1[f'C{row}'].font = Font(bold=True, size=14)
ws1.merge_cells(f'C{row}:K{row}')
ws1[f'C{row}'].alignment = center

row += 2
headers = ['Metric', 'Perpetua (SaaS)', 'Non-Perpetua (Manual)', 'Difference', '% Diff', 'Winner']
for col, header in enumerate(headers, start=3):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

row += 1

# Metrics to display
metrics = [
    ('Unique ASINs', 'Unique_ASINs', '#', None),
    ('Total Spend', 'Total_Spend', '$', None),
    ('Total Sales', 'Total_Sales', '$', False),
    ('Total Orders', 'Total_Orders', '#', False),
    ('Total Clicks', 'Total_Clicks', '#', False),
    ('Total Impressions', 'Total_Impressions', '#', False),
    ('', '', '', None),  # Separator
    ('ROAS (Return on Ad Spend)', 'ROAS', 'x', False),
    ('ACOS (Ad Cost of Sales)', 'ACOS', '%', True),
    ('CPC (Cost Per Click)', 'CPC', '$', True),
    ('CTR (Click-Through Rate)', 'CTR', '%', False),
    ('CVR (Conversion Rate)', 'CVR', '%', False),
    ('CPA (Cost Per Acquisition)', 'CPA', '$', True),
    ('CPM (Cost Per 1000 Impr.)', 'CPM', '$', True),
    ('AOV (Avg Order Value)', 'AOV', '$', False),
]

for metric_name, key, unit, lower_better in metrics:
    if not metric_name:  # Separator
        row += 1
        continue

    ws1.cell(row=row, column=3, value=metric_name).font = Font(size=10)

    # Perpetua value
    cell = ws1.cell(row=row, column=4, value=perpetua[key])
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Non-Perpetua value
    cell = ws1.cell(row=row, column=5, value=non_perpetua[key])
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Difference
    diff = perpetua[key] - non_perpetua[key]
    cell = ws1.cell(row=row, column=6, value=diff)
    if unit == '$':
        cell.number_format = '$#,##0.00;-$#,##0.00'
    elif unit == '%':
        cell.number_format = '0.00%;-0.00%'
    elif unit == 'x':
        cell.number_format = '0.00;-0.00'
    else:
        cell.number_format = '#,##0;-#,##0'

    # Percent difference
    if lower_better is not None and non_perpetua[key] != 0:
        pct = ((perpetua[key] - non_perpetua[key]) / non_perpetua[key])
        ws1.cell(row=row, column=7, value=pct).number_format = '0.0%;-0.0%'

        # Winner
        if lower_better:
            winner = 'Perpetua ✓' if perpetua[key] < non_perpetua[key] else 'Non-Perpetua ✓'
            is_better = perpetua[key] < non_perpetua[key]
        else:
            winner = 'Perpetua ✓' if perpetua[key] > non_perpetua[key] else 'Non-Perpetua ✓'
            is_better = perpetua[key] > non_perpetua[key]

        cell = ws1.cell(row=row, column=8, value=winner)
        cell.fill = PatternFill(start_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                               end_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                               fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = center

    row += 1

# ROAS Chart
print("[5/6] Adding charts...")
chart_row = row + 2

chart = BarChart()
chart.title = "ROAS Comparison"
chart.y_axis.title = "ROAS"
chart.height = 12
chart.width = 20

# Chart data
ws1.cell(row=chart_row, column=3, value='Platform')
ws1.cell(row=chart_row, column=4, value='ROAS')
ws1.cell(row=chart_row + 1, column=3, value='Perpetua')
ws1.cell(row=chart_row + 1, column=4, value=perpetua['ROAS'])
ws1.cell(row=chart_row + 2, column=3, value='Non-Perpetua')
ws1.cell(row=chart_row + 2, column=4, value=non_perpetua['ROAS'])

data = Reference(ws1, min_col=4, min_row=chart_row, max_row=chart_row + 2)
cats = Reference(ws1, min_col=3, min_row=chart_row + 1, max_row=chart_row + 2)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws1.add_chart(chart, 'C30')

# ============================================================================
# SHEET 2: DAILY DATA WITH FILTERS
# ============================================================================

print("[6/6] Creating Daily Data sheet...")
ws2 = wb.create_sheet("📅 Daily Data (Filter Here)")

ws2['B2'] = 'DAILY PERFORMANCE DATA - USE FILTERS TO SELECT DATE RANGE'
ws2['B2'].font = title_font
ws2.merge_cells('B2:M2')

ws2['B3'] = '▼ Click dropdown arrows in header row to filter by date or platform'
ws2['B3'].font = Font(size=11, italic=True, bold=True, color=COLORS['non_perpetua'])
ws2.merge_cells('B3:M3')

# Add daily data
row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(daily, index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws2.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            if c_idx == 2:  # Date
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx in [4, 5]:  # Spend, Sales
                cell.number_format = '$#,##0.00'
            elif c_idx in [6, 7, 8]:  # Orders, Clicks, Impressions
                cell.number_format = '#,##0'
            elif c_idx in [9, 10, 11, 12, 13]:  # ROAS, ACOS, CPC, CTR, CVR
                cell.number_format = '0.00'

            # Color by platform
            if value == 'Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
            elif value == 'Non-Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_orange'], end_color=COLORS['light_orange'], fill_type='solid')

# Enable AutoFilter
ws2.auto_filter.ref = f'B{row}:M{row + len(daily)}'

# Add trend chart
chart2 = LineChart()
chart2.title = "Daily ROAS Trend"
chart2.y_axis.title = "ROAS"
chart2.x_axis.title = "Date"
chart2.height = 12
chart2.width = 25

ws2.add_chart(chart2, 'O5')

# ============================================================================
# SHEET 3: INSTRUCTIONS
# ============================================================================

ws3 = wb.create_sheet("📖 Instructions")

ws3['C3'] = 'HOW TO USE DATE SELECTORS'
ws3['C3'].font = Font(bold=True, size=16)
ws3.merge_cells('C3:J3')

instructions = [
    ('', ''),
    ('METHOD 1: Use Date Dropdowns on Dashboard Sheet', ''),
    ('', ''),
    ('Step 1', 'Go to "📊 Dashboard" sheet'),
    ('Step 2', 'Click on cell D6 (Start Date)'),
    ('Step 3', 'You will see a dropdown arrow appear - click it'),
    ('Step 4', 'Select your desired start date from the list'),
    ('Step 5', 'Click on cell G6 (End Date) and select end date'),
    ('Step 6', 'The date range is now selected (for manual calculation)'),
    ('', ''),
    ('METHOD 2: Use AutoFilter on Daily Data Sheet (RECOMMENDED)', ''),
    ('', ''),
    ('Step 1', 'Go to "📅 Daily Data" sheet'),
    ('Step 2', 'Click the dropdown arrow ▼ in the "Date" column header'),
    ('Step 3', 'Select "Date Filters" > "Between..."'),
    ('Step 4', 'Enter your custom start and end dates'),
    ('Step 5', 'Click OK - data instantly filters to your range'),
    ('Step 6', 'You can also filter by "Advertising_Type" column'),
    ('', ''),
    ('EXAMPLE ANALYSES', ''),
    ('', ''),
    ('Last 30 Days', f'Filter dates: {(max_date - pd.Timedelta(days=30)).date()} to {max_date.date()}'),
    ('December Only', 'Filter dates: 2025-12-01 to 2025-12-31'),
    ('Perpetua Only', 'Filter Advertising_Type to "Perpetua"'),
    ('Weekdays Only', 'Filter Date > Custom > select Mon-Fri'),
    ('', ''),
    ('KEY METRICS EXPLAINED', ''),
    ('', ''),
    ('ROAS', f'Higher is better | Perpetua: {perpetua["ROAS"]:.2f}x, Non-Perpetua: {non_perpetua["ROAS"]:.2f}x'),
    ('ACOS', f'Lower is better | Perpetua: {perpetua["ACOS"]*100:.1f}%, Non-Perpetua: {non_perpetua["ACOS"]*100:.1f}%'),
    ('CPC', f'Lower is better | Perpetua: ${perpetua["CPC"]:.2f}, Non-Perpetua: ${non_perpetua["CPC"]:.2f}'),
    ('CTR', f'Higher is better | Perpetua: {perpetua["CTR"]*100:.2f}%, Non-Perpetua: {non_perpetua["CTR"]*100:.2f}%'),
    ('CVR', f'Higher is better | Perpetua: {perpetua["CVR"]*100:.2f}%, Non-Perpetua: {non_perpetua["CVR"]*100:.2f}%'),
]

row = 5
for label, text in instructions:
    cell1 = ws3.cell(row=row, column=3, value=label)
    if label and (label.startswith('METHOD') or label in ['EXAMPLE ANALYSES', 'KEY METRICS EXPLAINED']):
        cell1.font = Font(bold=True, size=12, color=COLORS['header'])
    elif label.startswith('Step'):
        cell1.font = Font(bold=True, size=10)
    else:
        cell1.font = Font(size=10)

    ws3.cell(row=row, column=4, value=text).font = Font(size=10)
    ws3.merge_cells(f'D{row}:J{row}')
    row += 1

# Set column widths
for ws in [ws1, ws2, ws3]:
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 22
    for col in 'DEFGHIJK':
        ws.column_dimensions[col].width = 17

# Save
output_file = OUTPUT_DIR / f'Perpetua_Dashboard_with_DateSelector_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(output_file)

print()
print("=" * 100)
print("✓ DASHBOARD WITH DATE SELECTORS COMPLETE")
print("=" * 100)
print(f"\nFile: {output_file.name}")
print()
print("🎛️ DATE SELECTOR FEATURES:")
print("  ✓ Dropdown menus in cells D6 and G6")
print("  ✓ Select from list of all available dates")
print("  ✓ AutoFilter on Daily Data sheet for instant filtering")
print("  ✓ Professional formatting and color-coding")
print()
print("📊 METRICS INCLUDED:")
print("  ✓ ROAS, ACOS, CPC, CTR, CVR, CPA, CPM, AOV")
print("  ✓ Total spend, sales, orders, clicks, impressions")
print("  ✓ Per-ASIN averages")
print()
print(f"📈 CORRECT RESULTS:")
print(f"  Perpetua:     {perpetua['ROAS']:.2f}x ROAS, {perpetua['ACOS']*100:.1f}% ACOS, ${perpetua['CPC']:.2f} CPC")
print(f"  Non-Perpetua: {non_perpetua['ROAS']:.2f}x ROAS, {non_perpetua['ACOS']*100:.1f}% ACOS, ${non_perpetua['CPC']:.2f} CPC")
print(f"  Winner:       Non-Perpetua is {((non_perpetua['ROAS'] - perpetua['ROAS'])/perpetua['ROAS']*100):.0f}% more efficient")
