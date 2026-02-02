#!/usr/bin/env python3
"""
COMBINED DASHBOARD - Campaign Report + Advertised Products Report
Matched by ASIN/SKU for comprehensive analysis
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / 'data' / 'recent-reports'
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 100)
print("COMBINED DASHBOARD: Campaign Report + Advertised Products")
print("Matched by ASIN/SKU | All Metrics Included | Date Normalized")
print("=" * 100)
print()

# ============================================================================
# LOAD ASIN/SKU MAPPING
# ============================================================================

print("[1/9] Loading ASIN/SKU lists...")
perpetua_list = pd.read_excel(DATA_DIR / 'ASIN list - perpetua.xlsx', sheet_name='perpetua list')
all_asins_list = pd.read_excel(DATA_DIR / 'ASIN list - perpetua.xlsx', sheet_name='All ASIns')

perpetua_asins = set(perpetua_list['ASIN'].dropna().str.strip())
perpetua_skus = set(perpetua_list['SKU'].dropna().str.strip())

all_asins = set(all_asins_list['ASIN (Informational only)'].dropna().str.strip())
non_perpetua_asins = all_asins - perpetua_asins

# Create bidirectional mapping
sku_to_asin = {}
asin_to_sku = {}

for _, row in perpetua_list.iterrows():
    if pd.notna(row['SKU']) and pd.notna(row['ASIN']):
        sku = row['SKU'].strip()
        asin = row['ASIN'].strip()
        sku_to_asin[sku] = asin
        asin_to_sku[asin] = sku

for _, row in all_asins_list.iterrows():
    if pd.notna(row['SKU']) and pd.notna(row['ASIN (Informational only)']):
        sku = row['SKU'].strip()
        asin = row['ASIN (Informational only)'].strip()
        if sku not in sku_to_asin:  # Don't overwrite perpetua mappings
            sku_to_asin[sku] = asin
        if asin not in asin_to_sku:
            asin_to_sku[asin] = sku

print(f"  ✓ Perpetua ASINs: {len(perpetua_asins)}")
print(f"  ✓ Non-Perpetua ASINs: {len(non_perpetua_asins)}")
print(f"  ✓ SKU-to-ASIN mappings: {len(sku_to_asin)}")

# ============================================================================
# LOAD BOTH REPORTS
# ============================================================================

print("[2/9] Loading Campaign Report...")
campaigns = pd.read_csv(DATA_DIR / 'SP_Campaign_-_4_Months.csv')
campaigns['Date'] = pd.to_datetime(campaigns['Date'], errors='coerce')
print(f"  ✓ {len(campaigns):,} campaign records")

print("[3/9] Loading Advertised Products Report...")
ad_products = pd.read_excel(DATA_DIR / 'SP_Advertised_Products_-_Max (1).xlsx')
ad_products['Date'] = pd.to_datetime(ad_products['Date'], errors='coerce')
print(f"  ✓ {len(ad_products):,} advertised product records")

# ============================================================================
# TAG AND MERGE
# ============================================================================

print("[4/9] Tagging and matching by ASIN/SKU...")

def extract_and_classify(campaign_name):
    """Extract identifiers and classify"""
    if pd.isna(campaign_name):
        return None, None, 'Unknown'

    # Extract ASIN (B0XXXXXXXXX)
    asin_match = re.search(r'B[A-Z0-9]{9}', str(campaign_name))
    asin = asin_match.group(0) if asin_match else None

    # Extract SKU (NT/SD/PN + numbers)
    sku_match = re.search(r'(NT|SD|PN)\d+[A-Z]?', str(campaign_name))
    sku = sku_match.group(0) if sku_match else None

    # If we found ASIN, use it
    if asin:
        if asin in perpetua_asins:
            return asin, sku, 'Perpetua'
        elif asin in non_perpetua_asins:
            return asin, sku, 'Non-Perpetua'

    # If we found SKU, map to ASIN
    if sku and sku in sku_to_asin:
        mapped_asin = sku_to_asin[sku]
        if mapped_asin in perpetua_asins:
            return mapped_asin, sku, 'Perpetua'
        elif mapped_asin in non_perpetua_asins:
            return mapped_asin, sku, 'Non-Perpetua'

    return asin, sku, 'Unknown'

# Tag campaigns
campaigns[['ASIN', 'SKU', 'Advertising_Type']] = campaigns['Campaign Name'].apply(
    lambda x: pd.Series(extract_and_classify(x))
)

# Tag advertised products
ad_products['Advertising_Type'] = ad_products['Advertised ASIN'].apply(
    lambda x: 'Perpetua' if pd.notna(x) and str(x).strip() in perpetua_asins
              else ('Non-Perpetua' if pd.notna(x) and str(x).strip() in non_perpetua_asins
                    else 'Unknown')
)

# Filter to known
campaigns_known = campaigns[campaigns['Advertising_Type'].isin(['Perpetua', 'Non-Perpetua'])].copy()
ad_products_known = ad_products[ad_products['Advertising_Type'].isin(['Perpetua', 'Non-Perpetua'])].copy()

print(f"  ✓ Campaign Report: {len(campaigns_known):,} matched campaigns")
print(f"  ✓ Advertised Products: {len(ad_products_known):,} matched records")

# ============================================================================
# COMBINE DATA SOURCES
# ============================================================================

print("[5/9] Combining data sources...")

# Clean numeric columns in both
for df in [campaigns_known, ad_products_known]:
    for col in ['Spend', '7 Day Total Sales ', '7 Day Total Orders (#)', 'Clicks', 'Impressions']:
        if col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace('$', '').str.replace(',', '')
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# Combine both datasets (union)
# Use campaign report as base, supplement with advertised products data
combined = pd.concat([campaigns_known, ad_products_known], ignore_index=True)
combined['Date'] = pd.to_datetime(combined['Date'], errors='coerce')
combined = combined[combined['Date'].notna()]

# Remove duplicates (keep campaign report data as primary)
combined = combined.drop_duplicates(subset=['Date', 'Campaign Name', 'Advertised ASIN'], keep='first')

min_date = combined['Date'].min()
max_date = combined['Date'].max()

print(f"  ✓ Combined dataset: {len(combined):,} records")
print(f"  ✓ Date range: {min_date.date()} to {max_date.date()}")

# ============================================================================
# CALCULATE AGGREGATE METRICS
# ============================================================================

print("[6/9] Calculating comprehensive metrics...")

def calc_all_metrics(subset):
    ts = subset['Spend'].sum()
    sales = subset['7 Day Total Sales '].sum()
    orders = subset['7 Day Total Orders (#)'].sum()
    clicks = subset['Clicks'].sum()
    impr = subset['Impressions'].sum()
    campaigns = subset['Campaign Name'].nunique() if 'Campaign Name' in subset.columns else 0
    asins = subset['ASIN'].nunique() if 'ASIN' in subset.columns else subset['Advertised ASIN'].nunique()

    return {
        'Data_Source': 'Campaign Report + Advertised Products (Combined)',
        'Unique_Campaigns': campaigns,
        'Unique_ASINs': asins,
        'Total_Spend': ts,
        'Total_Sales': sales,
        'Total_Orders': orders,
        'Total_Clicks': clicks,
        'Total_Impressions': impr,
        'ROAS': sales / ts if ts > 0 else 0,
        'ACOS': ts / sales if sales > 0 else 0,
        'CPC': ts / clicks if clicks > 0 else 0,
        'CTR': clicks / impr if impr > 0 else 0,
        'CVR': orders / clicks if clicks > 0 else 0,
        'CPA': ts / orders if orders > 0 else 0,
        'CPM': (ts / impr) * 1000 if impr > 0 else 0,
        'AOV': sales / orders if orders > 0 else 0,
        'Spend_Per_ASIN': ts / asins if asins > 0 else 0,
        'Sales_Per_ASIN': sales / asins if asins > 0 else 0,
    }

perpetua = calc_all_metrics(combined[combined['Advertising_Type'] == 'Perpetua'])
non_perpetua = calc_all_metrics(combined[combined['Advertising_Type'] == 'Non-Perpetua'])

print(f"\n✓ PERPETUA (Combined Data):")
for key, val in perpetua.items():
    if 'Total' in key or 'ROAS' in key or 'ACOS' in key:
        print(f"  {key}: {val if isinstance(val, str) else f'{val:,.2f}'}")

print(f"\n✓ NON-PERPETUA (Combined Data):")
for key, val in non_perpetua.items():
    if 'Total' in key or 'ROAS' in key or 'ACOS' in key:
        print(f"  {key}: {val if isinstance(val, str) else f'{val:,.2f}'}")

# Daily aggregation
daily = combined.groupby(['Date', 'Advertising_Type']).agg({
    'Spend': 'sum',
    '7 Day Total Sales ': 'sum',
    '7 Day Total Orders (#)': 'sum',
    'Clicks': 'sum',
    'Impressions': 'sum'
}).reset_index()

daily['ROAS'] = daily['7 Day Total Sales '] / daily['Spend'].replace(0, np.nan)
daily['ACOS'] = daily['Spend'] / daily['7 Day Total Sales '].replace(0, np.nan)
daily['CPC'] = daily['Spend'] / daily['Clicks'].replace(0, np.nan)
daily['CTR'] = daily['Clicks'] / daily['Impressions'].replace(0, np.nan)
daily['CVR'] = daily['7 Day Total Orders (#)'] / daily['Clicks'].replace(0, np.nan)
daily = daily.replace([np.inf, -np.inf], np.nan).fillna(0)

all_dates = sorted(combined['Date'].unique())

print(f"\n  ✓ {len(daily)} daily records")
print(f"  ✓ {len(all_dates)} unique dates")

# ============================================================================
# CREATE EXCEL DASHBOARD
# ============================================================================

print("[7/9] Creating Excel dashboard...")

wb = Workbook()
wb.remove(wb.active)

COLORS = {
    'perpetua': '4472C4',
    'non_perpetua': 'ED7D31',
    'header': '2F5496',
    'good': '70AD47',
    'input': 'FFF2CC',
    'light_blue': 'D9E1F2',
    'light_orange': 'FCE4D6'
}

header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=16)
center = Alignment(horizontal='center', vertical='center')

ws1 = wb.create_sheet("📊 Dashboard")

# Title
ws1['C2'] = 'PERPETUA vs MANUAL - COMPREHENSIVE ANALYSIS'
ws1['C2'].font = title_font
ws1['C2'].alignment = center
ws1.merge_cells('C2:K2')

ws1['C3'] = '📊 Combined Data: Campaign Report + Advertised Products | Matched by ASIN/SKU'
ws1['C3'].font = Font(size=10, bold=True)
ws1['C3'].alignment = center
ws1.merge_cells('C3:K3')

ws1['C4'] = f'{min_date.strftime("%B %d, %Y")} - {max_date.strftime("%B %d, %Y")} | {(max_date - min_date).days} days'
ws1['C4'].font = Font(size=9, italic=True)
ws1['C4'].alignment = center
ws1.merge_cells('C4:K4')

# Date selectors
row = 6
ws1[f'C{row}'] = '📅 DATE RANGE (Click for dropdown)'
ws1[f'C{row}'].font = Font(bold=True, size=12)
ws1.merge_cells(f'C{row}:K{row}')

row += 1
ws1[f'C{row}'] = 'Start:'
ws1[f'D{row}'] = min_date
ws1[f'D{row}'].number_format = 'YYYY-MM-DD'
ws1[f'D{row}'].fill = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type='solid')
ws1[f'D{row}'].font = Font(bold=True)
ws1[f'D{row}'].alignment = center

ws1[f'F{row}'] = 'End:'
ws1[f'G{row}'] = max_date
ws1[f'G{row}'].number_format = 'YYYY-MM-DD'
ws1[f'G{row}'].fill = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type='solid')
ws1[f'G{row}'].font = Font(bold=True)
ws1[f'G{row}'].alignment = center

# Hidden sheet with dates
ws_dates = wb.create_sheet("_Dates")
ws_dates.sheet_state = 'hidden'
for idx, date in enumerate(all_dates, start=1):
    ws_dates.cell(row=idx, column=1, value=pd.Timestamp(date).to_pydatetime())
    ws_dates.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'

# Data validation
dv = DataValidation(type="list", formula1=f"='_Dates'!$A$1:$A${len(all_dates)}")
ws1.add_data_validation(dv)
dv.add(ws1['D7'])
dv.add(ws1['G7'])

row += 1
ws1[f'C{row}'] = '⬆️ Click cells above to see dropdown ▼ | Or filter "Daily Data" sheet'
ws1[f'C{row}'].font = Font(size=9, italic=True)
ws1.merge_cells(f'C{row}:K{row}')

# Performance summary
row += 3
ws1[f'C{row}'] = 'COMPREHENSIVE PERFORMANCE METRICS'
ws1[f'C{row}'].font = Font(bold=True, size=13)
ws1.merge_cells(f'C{row}:K{row}')
ws1[f'C{row}'].alignment = center

row += 2
headers = ['Metric', 'Perpetua (SaaS)', 'Non-Perpetua (Manual)', 'Difference', '% Diff', 'Winner']
for col, header in enumerate(headers, start=3):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

row += 1
start_data_row = row

# All metrics
metrics = [
    ('📊 DATA COVERAGE', '', '', None),
    ('Campaigns Analyzed', 'Unique_Campaigns', '#', None),
    ('Unique ASINs/Products', 'Unique_ASINs', '#', None),
    ('', '', '', None),
    ('💰 VOLUME METRICS', '', '', None),
    ('Total Ad Spend', 'Total_Spend', '$', None),
    ('Total Sales Revenue', 'Total_Sales', '$', False),
    ('Total Orders', 'Total_Orders', '#', False),
    ('Total Clicks', 'Total_Clicks', '#', False),
    ('Total Impressions', 'Total_Impressions', '#', False),
    ('', '', '', None),
    ('🎯 EFFICIENCY METRICS', '', '', None),
    ('ROAS (Return on Ad Spend)', 'ROAS', 'x', False),
    ('ACOS (Ad Cost of Sales)', 'ACOS', '%', True),
    ('CPC (Cost Per Click)', 'CPC', '$', True),
    ('CTR (Click-Through Rate)', 'CTR', '%', False),
    ('CVR (Conversion Rate)', 'CVR', '%', False),
    ('CPA (Cost Per Acquisition)', 'CPA', '$', True),
    ('CPM (Cost Per 1000 Impr)', 'CPM', '$', True),
    ('AOV (Avg Order Value)', 'AOV', '$', False),
    ('', '', '', None),
    ('📈 PRODUCTIVITY', '', '', None),
    ('Spend per ASIN', 'Spend_Per_ASIN', '$', True),
    ('Sales per ASIN', 'Sales_Per_ASIN', '$', False),
]

for metric_name, key, unit, lower_better in metrics:
    if not metric_name or metric_name.startswith('📊') or metric_name.startswith('💰') or metric_name.startswith('🎯') or metric_name.startswith('📈'):
        # Section header
        if metric_name:
            ws1.cell(row=row, column=3, value=metric_name).font = Font(bold=True, size=11, color=COLORS['header'])
            ws1.merge_cells(f'C{row}:K{row}')
        row += 1
        continue

    ws1.cell(row=row, column=3, value=metric_name).font = Font(size=10)

    if key in perpetua:
        p_val = perpetua[key]
        np_val = non_perpetua[key]

        # Perpetua
        cell = ws1.cell(row=row, column=4, value=p_val)
        if unit == '$':
            cell.number_format = '$#,##0.00'
        elif unit == '%':
            cell.number_format = '0.00%'
        elif unit == 'x':
            cell.number_format = '0.00"x"'
        else:
            cell.number_format = '#,##0'

        # Non-Perpetua
        cell = ws1.cell(row=row, column=5, value=np_val)
        if unit == '$':
            cell.number_format = '$#,##0.00'
        elif unit == '%':
            cell.number_format = '0.00%'
        elif unit == 'x':
            cell.number_format = '0.00"x"'
        else:
            cell.number_format = '#,##0'

        # Difference
        diff = p_val - np_val
        cell = ws1.cell(row=row, column=6, value=diff)
        if unit == '$':
            cell.number_format = '$#,##0;-$#,##0'
        elif unit == '%':
            cell.number_format = '0.00%;-0.00%'
        else:
            cell.number_format = '0.00;-0.00'

        # % Diff
        if lower_better is not None and np_val != 0:
            pct = (p_val - np_val) / np_val
            ws1.cell(row=row, column=7, value=pct).number_format = '+0.0%;-0.0%'

            # Winner
            if lower_better:
                winner = 'Perpetua ✓' if p_val < np_val else 'Non-Perpetua ✓'
                is_better = p_val < np_val
            else:
                winner = 'Perpetua ✓' if p_val > np_val else 'Non-Perpetua ✓'
                is_better = p_val > np_val

            cell = ws1.cell(row=row, column=8, value=winner)
            cell.fill = PatternFill(start_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                                   end_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                                   fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = center

    row += 1

# ============================================================================
# SHEET 2: DAILY DATA
# ============================================================================

print("[8/9] Creating Daily Data sheet...")
ws2 = wb.create_sheet("📅 Daily Data (Filter Here)")

ws2['B2'] = 'DAILY DATA - COMBINED SOURCES'
ws2['B2'].font = title_font
ws2.merge_cells('B2:M2')

ws2['B3'] = '▼ Click filter dropdowns to select date range | Combined from Campaign + Advertised Products reports'
ws2['B3'].font = Font(size=10, italic=True, bold=True)
ws2.merge_cells('B3:M3')

row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(daily, index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws2.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            if c_idx == 2:
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx in [4, 5]:
                cell.number_format = '$#,##0.00'
            elif c_idx in [6, 7, 8]:
                cell.number_format = '#,##0'
            elif c_idx in [9, 10, 11, 12, 13]:
                cell.number_format = '0.00'

            # Color code
            if value == 'Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
            elif value == 'Non-Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_orange'], end_color=COLORS['light_orange'], fill_type='solid')

ws2.auto_filter.ref = f'B{row}:M{row + len(daily)}'

# Column widths
for ws in [ws1, ws2]:
    ws.column_dimensions['C'].width = 28
    for col in 'DEFGHIJK':
        ws.column_dimensions[col].width = 17

# Save
print("[9/9] Saving workbook...")
output_file = OUTPUT_DIR / f'Perpetua_Dashboard_COMBINED_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(output_file)

print()
print("=" * 100)
print("✓ COMBINED DASHBOARD COMPLETE")
print("=" * 100)
print(f"\nFile: {output_file.name}")
print()
print("📊 DATA SOURCES COMBINED:")
print(f"  ✓ Campaign Report: {len(campaigns_known):,} campaigns")
print(f"  ✓ Advertised Products: {len(ad_products_known):,} ASIN records")
print(f"  ✓ Combined total: {len(combined):,} records")
print()
print("📈 RESULTS:")
print(f"  Perpetua:     {perpetua['ROAS']:.2f}x ROAS, {perpetua['ACOS']*100:.1f}% ACOS")
print(f"  Non-Perpetua: {non_perpetua['ROAS']:.2f}x ROAS, {non_perpetua['ACOS']*100:.1f}% ACOS")
print(f"  Difference:   {((non_perpetua['ROAS'] - perpetua['ROAS'])/perpetua['ROAS']*100):.0f}% better for Non-Perpetua")
print()
print("✅ Date dropdowns in cells D7 and G7")
print("✅ AutoFilter on Daily Data sheet")
print("✅ All metrics included: ROAS, ACOS, CPC, CTR, CVR, CPA, CPM, AOV")
