#!/usr/bin/env python3
"""
Generate Dashboard from Campaign Report (PRIMARY SOURCE)
Uses SP_Campaign_-_4_Months.csv as the basis for all metrics
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / 'data' / 'recent-reports'
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 100)
print("DASHBOARD GENERATION FROM CAMPAIGN REPORT (PRIMARY DATA SOURCE)")
print("=" * 100)
print()

# ============================================================================
# LOAD CAMPAIGN REPORT
# ============================================================================

print("[1/8] Loading Campaign Report...")
campaigns = pd.read_csv(DATA_DIR / 'SP_Campaign_-_4_Months.csv')
print(f"  ✓ Loaded {len(campaigns):,} campaign records")
print(f"  ✓ Unique campaigns: {campaigns['Campaign Name'].nunique():,}")
print(f"  ✓ Date range: {campaigns['Date'].min()} to {campaigns['Date'].max()}")

# ============================================================================
# LOAD ASIN LISTS FOR TAGGING
# ============================================================================

print("[2/8] Loading ASIN lists for Perpetua tagging...")
perpetua_df = pd.read_excel(DATA_DIR / 'ASIN list - perpetua.xlsx', sheet_name='perpetua list')
all_asins_df = pd.read_excel(DATA_DIR / 'ASIN list - perpetua.xlsx', sheet_name='All ASIns')

# Create lookup dictionaries (ASIN -> SKU and SKU -> ASIN)
perpetua_asins = set(perpetua_df['ASIN'].dropna().str.strip())
perpetua_skus = set(perpetua_df['SKU'].dropna().str.strip())

all_asins = set(all_asins_df['ASIN (Informational only)'].dropna().str.strip())
non_perpetua_asins = all_asins - perpetua_asins

# Create SKU to ASIN mapping
sku_to_asin = {}
for _, row in perpetua_df.iterrows():
    if pd.notna(row['SKU']) and pd.notna(row['ASIN']):
        sku_to_asin[row['SKU'].strip()] = row['ASIN'].strip()

for _, row in all_asins_df.iterrows():
    if pd.notna(row['SKU']) and pd.notna(row['ASIN (Informational only)']):
        sku_to_asin[row['SKU'].strip()] = row['ASIN (Informational only)'].strip()

print(f"  ✓ Perpetua ASINs: {len(perpetua_asins)}")
print(f"  ✓ Perpetua SKUs: {len(perpetua_skus)}")
print(f"  ✓ Non-Perpetua ASINs: {len(non_perpetua_asins)}")
print(f"  ✓ SKU-to-ASIN mappings: {len(sku_to_asin)}")

# ============================================================================
# TAG CAMPAIGNS AS PERPETUA OR NON-PERPETUA
# ============================================================================

print("[3/8] Tagging campaigns as Perpetua vs Non-Perpetua...")

def extract_identifiers(campaign_name):
    """Extract ASIN and SKU from campaign name"""
    if pd.isna(campaign_name):
        return None, None

    # Look for ASIN pattern (B0XXXXXXXXX)
    asin_match = re.search(r'B[A-Z0-9]{9}', str(campaign_name))
    asin = asin_match.group(0) if asin_match else None

    # Look for SKU patterns (NT12780A, SD1511, etc.)
    sku_match = re.search(r'(NT|SD|PN)\d+[A-Z]?', str(campaign_name))
    sku = sku_match.group(0) if sku_match else None

    return asin, sku

def classify_campaign(campaign_name):
    """Classify campaign as Perpetua or Non-Perpetua"""
    asin, sku = extract_identifiers(campaign_name)

    # Try ASIN first
    if asin:
        if asin in perpetua_asins:
            return 'Perpetua', asin, sku
        elif asin in non_perpetua_asins:
            return 'Non-Perpetua', asin, sku

    # Try SKU
    if sku:
        if sku in perpetua_skus:
            return 'Perpetua', sku_to_asin.get(sku), sku
        # Check if SKU maps to a non-Perpetua ASIN
        if sku in sku_to_asin:
            mapped_asin = sku_to_asin[sku]
            if mapped_asin in non_perpetua_asins:
                return 'Non-Perpetua', mapped_asin, sku

    return 'Unknown', asin, sku

# Apply classification
campaigns[['Advertising_Type', 'ASIN', 'SKU']] = campaigns['Campaign Name'].apply(
    lambda x: pd.Series(classify_campaign(x))
)

# Distribution
type_counts = campaigns['Advertising_Type'].value_counts()
print(f"\n  Campaign Classification:")
for ad_type, count in type_counts.items():
    pct = count / len(campaigns) * 100
    print(f"    {ad_type:15s}: {count:7,} ({pct:5.1f}%)")

# Filter to known campaigns
known = campaigns[campaigns['Advertising_Type'].isin(['Perpetua', 'Non-Perpetua'])].copy()
print(f"\n  ✓ Analyzing {len(known):,} campaigns with known classification")

# ============================================================================
# CLEAN AND PREPARE METRICS
# ============================================================================

print("[4/8] Cleaning campaign data...")

# Convert dates
known['Date'] = pd.to_datetime(known['Date'], errors='coerce')
known = known[known['Date'].notna()]

# Clean numeric columns
numeric_cols = ['Spend', '7 Day Total Sales ', '7 Day Total Orders (#)',
                'Clicks', 'Impressions', 'Budget Amount']

for col in numeric_cols:
    if col in known.columns:
        if known[col].dtype == 'object':
            known[col] = known[col].astype(str).str.replace('$', '').str.replace(',', '')
        known[col] = pd.to_numeric(known[col], errors='coerce').fillna(0)

# Clean percentage columns
for col in ['Click-Thru Rate (CTR)', 'Total Advertising Cost of Sales (ACOS) ']:
    if col in known.columns:
        if known[col].dtype == 'object':
            known[col] = known[col].astype(str).str.replace('%', '')
        known[col] = pd.to_numeric(known[col], errors='coerce').fillna(0) / 100

# Clean ROAS
if 'Total Return on Advertising Spend (ROAS)' in known.columns:
    known['ROAS'] = pd.to_numeric(known['Total Return on Advertising Spend (ROAS)'], errors='coerce').fillna(0)

# Calculate derived metrics
known['CPC_calc'] = np.where(known['Clicks'] > 0, known['Spend'] / known['Clicks'], 0)
known['CTR_calc'] = np.where(known['Impressions'] > 0, known['Clicks'] / known['Impressions'], 0)
known['CVR'] = np.where(known['Clicks'] > 0, known['7 Day Total Orders (#)'] / known['Clicks'], 0)
known['CPA'] = np.where(known['7 Day Total Orders (#)'] > 0,
                        known['Spend'] / known['7 Day Total Orders (#)'], 0)

min_date = known['Date'].min()
max_date = known['Date'].max()

print(f"  ✓ Cleaned data: {min_date.date()} to {max_date.date()}")
print(f"  ✓ {len(known):,} campaign records ready for analysis")

# ============================================================================
# AGGREGATE METRICS BY PLATFORM
# ============================================================================

print("[5/8] Calculating aggregate metrics from campaign data...")

def calc_metrics(subset):
    total_spend = subset['Spend'].sum()
    total_sales = subset['7 Day Total Sales '].sum()
    total_orders = subset['7 Day Total Orders (#)'].sum()
    total_clicks = subset['Clicks'].sum()
    total_impressions = subset['Impressions'].sum()
    unique_campaigns = subset['Campaign Name'].nunique()
    unique_asins = subset['ASIN'].nunique()

    return {
        'Total_Spend': total_spend,
        'Total_Sales': total_sales,
        'Total_Orders': total_orders,
        'Total_Clicks': total_clicks,
        'Total_Impressions': total_impressions,
        'Unique_Campaigns': unique_campaigns,
        'Unique_ASINs': unique_asins,
        'ROAS': total_sales / total_spend if total_spend > 0 else 0,
        'ACOS': total_spend / total_sales if total_sales > 0 else 0,
        'CPC': total_spend / total_clicks if total_clicks > 0 else 0,
        'CTR': total_clicks / total_impressions if total_impressions > 0 else 0,
        'CVR': total_orders / total_clicks if total_clicks > 0 else 0,
        'CPA': total_spend / total_orders if total_orders > 0 else 0,
        'CPM': (total_spend / total_impressions) * 1000 if total_impressions > 0 else 0,
        'AOV': total_sales / total_orders if total_orders > 0 else 0,
    }

perpetua = calc_metrics(known[known['Advertising_Type'] == 'Perpetua'])
non_perpetua = calc_metrics(known[known['Advertising_Type'] == 'Non-Perpetua'])

print(f"\nPERPETUA (from Campaign Report):")
print(f"  Campaigns: {perpetua['Unique_Campaigns']:,}")
print(f"  ASINs: {perpetua['Unique_ASINs']:,}")
print(f"  Spend: ${perpetua['Total_Spend']:,.2f}")
print(f"  Sales: ${perpetua['Total_Sales']:,.2f}")
print(f"  ROAS: {perpetua['ROAS']:.2f}x")
print(f"  ACOS: {perpetua['ACOS']*100:.1f}%")

print(f"\nNON-PERPETUA (from Campaign Report):")
print(f"  Campaigns: {non_perpetua['Unique_Campaigns']:,}")
print(f"  ASINs: {non_perpetua['Unique_ASINs']:,}")
print(f"  Spend: ${non_perpetua['Total_Spend']:,.2f}")
print(f"  Sales: ${non_perpetua['Total_Sales']:,.2f}")
print(f"  ROAS: {non_perpetua['ROAS']:.2f}x")
print(f"  ACOS: {non_perpetua['ACOS']*100:.1f}%")

# Daily aggregation for time series
daily = known.groupby(['Date', 'Advertising_Type']).agg({
    'Spend': 'sum',
    '7 Day Total Sales ': 'sum',
    '7 Day Total Orders (#)': 'sum',
    'Clicks': 'sum',
    'Impressions': 'sum'
}).reset_index()

daily['ROAS'] = daily['7 Day Total Sales '] / daily['Spend'].replace(0, np.nan)
daily['ACOS'] = daily['Spend'] / daily['7 Day Total Sales '].replace(0, np.nan)
daily['CPC'] = daily['Spend'] / daily['Clicks'].replace(0, np.nan)
daily['CTR'] = daily['Clicks'] / daily['Impressions'].replace(0, np.nan)
daily['CVR'] = daily['7 Day Total Orders (#)'] / daily['Clicks'].replace(0, np.nan)
daily = daily.replace([np.inf, -np.inf], np.nan).fillna(0)

all_dates = sorted(known['Date'].unique())
print(f"\n  ✓ {len(daily)} daily records prepared")
print(f"  ✓ {len(all_dates)} unique dates for selector")

# ============================================================================
# CREATE EXCEL WORKBOOK
# ============================================================================

print("[6/8] Creating Excel workbook...")

wb = Workbook()
wb.remove(wb.active)

COLORS = {
    'perpetua': '4472C4',
    'non_perpetua': 'ED7D31',
    'header': '2F5496',
    'good': '70AD47',
    'bad': 'C5504B',
    'input': 'FFF2CC',
    'light_blue': 'D9E1F2',
    'light_orange': 'FCE4D6'
}

header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=16)
center = Alignment(horizontal='center', vertical='center')

# ============================================================================
# SHEET 1: EXECUTIVE DASHBOARD
# ============================================================================

print("[7/8] Creating Executive Dashboard...")
ws1 = wb.create_sheet("📊 Dashboard")

# Title
ws1['C2'] = 'CAMPAIGN PERFORMANCE: PERPETUA vs MANUAL'
ws1['C2'].font = title_font
ws1['C2'].alignment = center
ws1.merge_cells('C2:K2')

ws1['C3'] = f'Based on Campaign Report | {min_date.strftime("%b %d, %Y")} - {max_date.strftime("%b %d, %Y")}'
ws1['C3'].font = Font(size=11, italic=True)
ws1['C3'].alignment = center
ws1.merge_cells('C3:K3')

# Data source indicator
ws1['C4'] = f'📄 Data Source: SP_Campaign_-_4_Months.csv ({len(known):,} campaigns analyzed)'
ws1['C4'].font = Font(size=9, color='666666')
ws1.merge_cells('C4:K4')
ws1['C4'].alignment = center

# DATE SELECTORS
row = 6
ws1[f'C{row}'] = '📅 DATE RANGE SELECTOR'
ws1[f'C{row}'].font = Font(bold=True, size=13)
ws1.merge_cells(f'C{row}:K{row}')

row += 1
ws1[f'C{row}'] = 'Start Date:'
ws1[f'D{row}'] = min_date
ws1[f'D{row}'].number_format = 'YYYY-MM-DD'
ws1[f'D{row}'].fill = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type='solid')
ws1[f'D{row}'].font = Font(bold=True, size=11)
ws1[f'D{row}'].alignment = center

ws1[f'F{row}'] = 'End Date:'
ws1[f'G{row}'] = max_date
ws1[f'G{row}'].number_format = 'YYYY-MM-DD'
ws1[f'G{row}'].fill = PatternFill(start_color=COLORS['input'], end_color=COLORS['input'], fill_type='solid')
ws1[f'G{row}'].font = Font(bold=True, size=11)
ws1[f'G{row}'].alignment = center

# Create hidden sheet for date list
ws_dates = wb.create_sheet("_Dates")
ws_dates.sheet_state = 'hidden'
for idx, date in enumerate(all_dates, start=1):
    ws_dates.cell(row=idx, column=1, value=pd.Timestamp(date).to_pydatetime())
    ws_dates.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'

# Add data validation dropdowns
dv_start = DataValidation(type="list", formula1=f"='_Dates'!$A$1:$A${len(all_dates)}")
ws1.add_data_validation(dv_start)
dv_start.add(ws1['D7'])

dv_end = DataValidation(type="list", formula1=f"='_Dates'!$A$1:$A${len(all_dates)}")
ws1.add_data_validation(dv_end)
dv_end.add(ws1['G7'])

row += 1
ws1[f'C{row}'] = '👆 Click D7 and G7 to see date dropdown menus | Or use "Daily Data" sheet filters'
ws1[f'C{row}'].font = Font(size=9, italic=True)
ws1.merge_cells(f'C{row}:K{row}')

# METRICS TABLE
row += 3
ws1[f'C{row}'] = 'ALL PERFORMANCE METRICS'
ws1[f'C{row}'].font = Font(bold=True, size=14)
ws1.merge_cells(f'C{row}:K{row}')
ws1[f'C{row}'].alignment = center

row += 2
headers = ['Metric', 'Perpetua (SaaS)', 'Non-Perpetua', 'Difference', '% Diff', 'Winner']
for col, header in enumerate(headers, start=3):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

row += 1

metrics = [
    ('Campaigns Analyzed', 'Unique_Campaigns', '#', None),
    ('Unique ASINs', 'Unique_ASINs', '#', None),
    ('Total Spend', 'Total_Spend', '$', None),
    ('Total Sales (7-day attr.)', 'Total_Sales', '$', False),
    ('Total Orders', 'Total_Orders', '#', False),
    ('Total Clicks', 'Total_Clicks', '#', False),
    ('Total Impressions', 'Total_Impressions', '#', False),
    ('', '', '', None),
    ('ROAS', 'ROAS', 'x', False),
    ('ACOS', 'ACOS', '%', True),
    ('CPC', 'CPC', '$', True),
    ('CTR', 'CTR', '%', False),
    ('CVR', 'CVR', '%', False),
    ('CPA', 'CPA', '$', True),
    ('CPM', 'CPM', '$', True),
    ('AOV', 'AOV', '$', False),
]

for metric_name, key, unit, lower_better in metrics:
    if not metric_name:
        row += 1
        continue

    ws1.cell(row=row, column=3, value=metric_name)

    # Values
    p_val = perpetua[key]
    np_val = non_perpetua[key]

    cell = ws1.cell(row=row, column=4, value=p_val)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    cell = ws1.cell(row=row, column=5, value=np_val)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Difference
    diff = p_val - np_val
    cell = ws1.cell(row=row, column=6, value=diff)
    if unit == '$':
        cell.number_format = '$#,##0;-$#,##0'
    elif unit == '%':
        cell.number_format = '0.00%;-0.00%'
    elif unit == 'x':
        cell.number_format = '0.00;-0.00'
    else:
        cell.number_format = '#,##0;-#,##0'

    # % Diff and Winner
    if lower_better is not None and np_val != 0:
        pct = ((p_val - np_val) / np_val)
        ws1.cell(row=row, column=7, value=pct).number_format = '0.0%;-0.0%'

        if lower_better:
            winner = 'Perpetua ✓' if p_val < np_val else 'Non-Perpetua ✓'
            is_better = p_val < np_val
        else:
            winner = 'Perpetua ✓' if p_val > np_val else 'Non-Perpetua ✓'
            is_better = p_val > np_val

        cell = ws1.cell(row=row, column=8, value=winner)
        cell.fill = PatternFill(start_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                               end_color=COLORS['perpetua'] if is_better else COLORS['non_perpetua'],
                               fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = center

    row += 1

# ============================================================================
# SHEET 2: DAILY DATA
# ============================================================================

print("[8/8] Creating Daily Data sheet with filters...")
ws2 = wb.create_sheet("📅 Daily Data")

ws2['B2'] = 'DAILY CAMPAIGN PERFORMANCE - FILTER BY DATE HERE'
ws2['B2'].font = title_font
ws2.merge_cells('B2:M2')

ws2['B3'] = '▼ Click dropdown arrows to filter | Data from Campaign Report'
ws2['B3'].font = Font(size=10, italic=True, bold=True)
ws2.merge_cells('B3:M3')

row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(daily, index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws2.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            if c_idx == 2:
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx in [4, 5]:
                cell.number_format = '$#,##0.00'
            elif c_idx in [6, 7, 8]:
                cell.number_format = '#,##0'
            elif c_idx in [9, 10, 11, 12, 13]:
                cell.number_format = '0.00'

ws2.auto_filter.ref = f'B{row}:M{row + len(daily)}'

# Widths
for ws in [ws1, ws2]:
    ws.column_dimensions['C'].width = 25
    for col in 'DEFGHIJK':
        ws.column_dimensions[col].width = 16

# Save
output_file = OUTPUT_DIR / f'Campaign_Report_Dashboard_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(output_file)

print()
print("=" * 100)
print("✓ CAMPAIGN REPORT DASHBOARD COMPLETE")
print("=" * 100)
print(f"\nFile: {output_file.name}")
print()
print(f"📊 BASIS: Campaign Report (SP_Campaign_-_4_Months.csv)")
print(f"  ✓ {len(known):,} campaign records analyzed")
print(f"  ✓ Date range: {min_date.date()} to {max_date.date()}")
print(f"  ✓ Perpetua: {perpetua['Unique_Campaigns']:,} campaigns, {perpetua['Unique_ASINs']} ASINs")
print(f"  ✓ Non-Perpetua: {non_perpetua['Unique_Campaigns']:,} campaigns, {non_perpetua['Unique_ASINs']} ASINs")
print()
print("✅ Date selectors in cells D7 and G7 (dropdown menus)")
print("✅ All metrics: ROAS, ACOS, CPC, CTR, CVR, CPA, CPM, AOV")
print("✅ AutoFilter enabled on Daily Data sheet")
