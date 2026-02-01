#!/usr/bin/env python3
"""
Generate Excel Dashboard with multiple sheets and visualizations
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
BASE_DIR = Path(__file__).parent.parent
AGG_DIR = BASE_DIR / 'data' / 'aggregated'
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 80)
print("GENERATING EXCEL DASHBOARD")
print("=" * 80)
print()

# Load data
print("[1/5] Loading processed data...")
with open(AGG_DIR / 'asin_level_comparison.json', 'r') as f:
    analysis = json.load(f)

comparison_df = pd.read_csv(AGG_DIR / 'asin_comparison_full.csv')
processed_campaigns = pd.read_csv(PROCESSED_DIR / 'advertised_products_processed.csv', low_memory=False)
print("  ✓ Data loaded")

# Extract metrics
perpetua = analysis['perpetua_metrics']
non_perpetua = analysis['non_perpetua_metrics']

# Create Excel file
excel_file = OUTPUT_DIR / f'Perpetua_Performance_Dashboard_{datetime.now().strftime("%Y%m%d")}.xlsx'

print("[2/5] Creating Excel sheets...")

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # Sheet 1: Executive Summary
    summary_data = {
        'Metric': [
            'Report Generated',
            'Analysis Period',
            '',
            'Perpetua ASINs Analyzed',
            'Non-Perpetua ASINs Analyzed',
            'Total ASINs',
            '',
            '=== PERPETUA PERFORMANCE ===',
            'Total Spend',
            'Total Sales',
            'Total Orders',
            'ROAS',
            'ACOS',
            'Avg CPC',
            'Conversion Rate',
            '',
            '=== NON-PERPETUA PERFORMANCE ===',
            'Total Spend',
            'Total Sales',
            'Total Orders',
            'ROAS',
            'ACOS',
            'Avg CPC',
            'Conversion Rate',
            '',
            '=== COMPARISON ===',
            'ROAS Difference',
            'ACOS Difference',
            'CPC Difference',
            'Conversion Rate Difference'
        ],
        'Value': [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Last 4 Months',
            '',
            analysis['perpetua_asins_analyzed'],
            analysis['non_perpetua_asins_analyzed'],
            analysis['perpetua_asins_count'] + analysis['non_perpetua_asins_count'],
            '',
            '',
            f"${perpetua['Total_Spend']:,.2f}",
            f"${perpetua['Total_Sales']:,.2f}",
            f"{perpetua['Total_Orders']:,.0f}",
            f"{perpetua['ROAS']:.2f}x",
            f"{perpetua['ACOS']*100:.2f}%",
            f"${perpetua['Avg_CPC']:.2f}",
            f"{perpetua['Avg_CVR']*100:.2f}%",
            '',
            '',
            f"${non_perpetua['Total_Spend']:,.2f}",
            f"${non_perpetua['Total_Sales']:,.2f}",
            f"{non_perpetua['Total_Orders']:,.0f}",
            f"{non_perpetua['ROAS']:.2f}x",
            f"{non_perpetua['ACOS']*100:.2f}%",
            f"${non_perpetua['Avg_CPC']:.2f}",
            f"{non_perpetua['Avg_CVR']*100:.2f}%",
            '',
            '',
            f"{perpetua['ROAS'] - non_perpetua['ROAS']:.2f}x",
            f"{(perpetua['ACOS'] - non_perpetua['ACOS'])*100:.2f}pp",
            f"${perpetua['Avg_CPC'] - non_perpetua['Avg_CPC']:.2f}",
            f"{(perpetua['Avg_CVR'] - non_perpetua['Avg_CVR'])*100:.2f}pp"
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)

    # Sheet 2: Detailed Comparison
    comparison_data = {
        'Metric': [
            'Total Spend',
            'Total Sales',
            'Total Orders',
            'Total Units',
            'Total Impressions',
            'Total Clicks',
            'Unique ASINs',
            'Avg ACOS',
            'Avg ROAS',
            'Avg CPC',
            'Avg Conversion Rate',
            'CTR',
            'Spend per ASIN',
            'Sales per ASIN',
            'Orders per ASIN'
        ],
        'Perpetua': [
            f"${perpetua['Total_Spend']:,.2f}",
            f"${perpetua['Total_Sales']:,.2f}",
            f"{perpetua['Total_Orders']:,.0f}",
            f"{perpetua['Total_Units']:,.0f}",
            f"{perpetua['Total_Impressions']:,.0f}",
            f"{perpetua['Total_Clicks']:,.0f}",
            f"{perpetua['Unique_ASINs']:.0f}",
            f"{perpetua['ACOS']*100:.2f}%",
            f"{perpetua['ROAS']:.2f}x",
            f"${perpetua['Avg_CPC']:.2f}",
            f"{perpetua['Avg_CVR']*100:.2f}%",
            f"{perpetua['CTR']*100:.2f}%",
            f"${perpetua['Spend_Per_ASIN']:,.2f}",
            f"${perpetua['Sales_Per_ASIN']:,.2f}",
            f"{perpetua['Orders_Per_ASIN']:,.0f}"
        ],
        'Non-Perpetua': [
            f"${non_perpetua['Total_Spend']:,.2f}",
            f"${non_perpetua['Total_Sales']:,.2f}",
            f"{non_perpetua['Total_Orders']:,.0f}",
            f"{non_perpetua['Total_Units']:,.0f}",
            f"{non_perpetua['Total_Impressions']:,.0f}",
            f"{non_perpetua['Total_Clicks']:,.0f}",
            f"{non_perpetua['Unique_ASINs']:.0f}",
            f"{non_perpetua['ACOS']*100:.2f}%",
            f"{non_perpetua['ROAS']:.2f}x",
            f"${non_perpetua['Avg_CPC']:.2f}",
            f"{non_perpetua['Avg_CVR']*100:.2f}%",
            f"{non_perpetua['CTR']*100:.2f}%",
            f"${non_perpetua['Spend_Per_ASIN']:,.2f}",
            f"${non_perpetua['Sales_Per_ASIN']:,.2f}",
            f"{non_perpetua['Orders_Per_ASIN']:,.0f}"
        ],
        'Winner': [
            'N/A',
            'Perpetua' if perpetua['Total_Sales'] > non_perpetua['Total_Sales'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Total_Orders'] > non_perpetua['Total_Orders'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Total_Units'] > non_perpetua['Total_Units'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Total_Impressions'] > non_perpetua['Total_Impressions'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Total_Clicks'] > non_perpetua['Total_Clicks'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Unique_ASINs'] > non_perpetua['Unique_ASINs'] else 'Non-Perpetua',
            'Perpetua' if perpetua['ACOS'] < non_perpetua['ACOS'] else 'Non-Perpetua',
            'Perpetua' if perpetua['ROAS'] > non_perpetua['ROAS'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Avg_CPC'] < non_perpetua['Avg_CPC'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Avg_CVR'] > non_perpetua['Avg_CVR'] else 'Non-Perpetua',
            'Perpetua' if perpetua['CTR'] > non_perpetua['CTR'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Spend_Per_ASIN'] < non_perpetua['Spend_Per_ASIN'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Sales_Per_ASIN'] > non_perpetua['Sales_Per_ASIN'] else 'Non-Perpetua',
            'Perpetua' if perpetua['Orders_Per_ASIN'] > non_perpetua['Orders_Per_ASIN'] else 'Non-Perpetua'
        ]
    }

    comp_df = pd.DataFrame(comparison_data)
    comp_df.to_excel(writer, sheet_name='Detailed Comparison', index=False)

    # Sheet 3: ASIN-Level Data (sample of top performers)
    print("  ✓ Creating ASIN-level summary...")

    # Aggregate by ASIN
    asin_summary = processed_campaigns.groupby(['Advertised ASIN', 'Advertising_Type']).agg({
        'Spend': 'sum',
        '7 Day Total Sales ': 'sum',
        '7 Day Total Orders (#)': 'sum',
        'Clicks': 'sum',
        'Impressions': 'sum'
    }).reset_index()

    # Calculate metrics
    asin_summary['ROAS'] = asin_summary['7 Day Total Sales '] / asin_summary['Spend']
    asin_summary['ACOS'] = asin_summary['Spend'] / asin_summary['7 Day Total Sales ']
    asin_summary['CVR'] = asin_summary['7 Day Total Orders (#)'] / asin_summary['Clicks']

    # Sort by sales
    asin_summary = asin_summary.sort_values('7 Day Total Sales ', ascending=False)

    # Take top 100 for Excel
    asin_summary.head(100).to_excel(writer, sheet_name='Top 100 ASINs', index=False)

    # Sheet 4: Monthly Trends
    print("  ✓ Creating monthly trend data...")

    # Convert date column
    processed_campaigns['Date'] = pd.to_datetime(processed_campaigns['Date'])
    processed_campaigns['Month'] = processed_campaigns['Date'].dt.to_period('M')

    monthly_summary = processed_campaigns.groupby(['Month', 'Advertising_Type']).agg({
        'Spend': 'sum',
        '7 Day Total Sales ': 'sum',
        '7 Day Total Orders (#)': 'sum',
        'Clicks': 'sum'
    }).reset_index()

    monthly_summary['ROAS'] = monthly_summary['7 Day Total Sales '] / monthly_summary['Spend']
    monthly_summary['ACOS'] = monthly_summary['Spend'] / monthly_summary['7 Day Total Sales ']

    monthly_summary.to_excel(writer, sheet_name='Monthly Trends', index=False)

    # Sheet 5: Recommendations
    recommendations = {
        'Priority': ['HIGH', 'HIGH', 'MEDIUM', 'MEDIUM', 'LOW'],
        'Category': [
            'Efficiency',
            'Strategy',
            'Optimization',
            'Expansion',
            'Monitoring'
        ],
        'Recommendation': [
            f"Non-Perpetua shows {((non_perpetua['ROAS'] - perpetua['ROAS']) / perpetua['ROAS'] * 100):.1f}% better ROAS - investigate and apply strategies",
            'Focus automation on high-volume products while preserving non-Perpetua efficiency',
            'Expand negative keyword management to reduce wasted spend',
            'Consider migrating high-performing non-Perpetua ASINs to Perpetua management',
            'Implement automated dashboard refresh to track improvements over time'
        ],
        'Expected Impact': [
            f"Potential ROAS improvement of {((non_perpetua['ROAS'] - perpetua['ROAS']) / perpetua['ROAS'] * 100):.0f}%",
            'Maintain scale while improving efficiency',
            'Reduce ACOS by 5-10%',
            'Increase automated coverage from 227 to 300+ ASINs',
            'Faster identification of optimization opportunities'
        ]
    }

    rec_df = pd.DataFrame(recommendations)
    rec_df.to_excel(writer, sheet_name='Recommendations', index=False)

print("  ✓ Excel file created")

# Format the workbook
print("[3/5] Applying formatting...")

wb = load_workbook(excel_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=14)
center_alignment = Alignment(horizontal="center", vertical="center")
currency_format = '$#,##0.00'
percent_format = '0.00%'
number_format = '#,##0'

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Set column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Format headers (first row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

print("  ✓ Formatting applied")

# Save formatted workbook
print("[4/5] Saving workbook...")
wb.save(excel_file)
print(f"  ✓ Saved: {excel_file}")

print()
print("[5/5] Creating usage instructions...")

instructions = f"""
# Excel Dashboard Usage Instructions

## File Location
{excel_file}

## Sheets Overview

1. **Executive Summary**
   - High-level metrics at a glance
   - Perpetua vs Non-Perpetua comparison
   - Key performance indicators

2. **Detailed Comparison**
   - Side-by-side metrics
   - Winner identification for each metric
   - Comprehensive performance analysis

3. **Top 100 ASINs**
   - ASIN-level performance data
   - Sorted by total sales
   - Identify star performers and opportunities

4. **Monthly Trends**
   - Month-over-month performance
   - Track improvements over time
   - Seasonal pattern identification

5. **Recommendations**
   - Prioritized action items
   - Expected impact assessment
   - Strategic next steps

## How to Use

1. **Review Executive Summary** first for high-level insights
2. **Dive into Detailed Comparison** for metric-by-metric analysis
3. **Examine Top 100 ASINs** to identify specific optimization targets
4. **Check Monthly Trends** to track progress over time
5. **Follow Recommendations** for actionable next steps

## Refreshing Data

To update this dashboard with new data:
```bash
cd /mnt/c/Users/Krell/Documents/Imps/gits/Perpetua-Report
python scripts/refresh_reports.py
```

This will:
1. Re-process all data files
2. Re-generate comparison analysis
3. Create new Excel dashboard with updated date
4. Preserve historical dashboards for comparison

## Tips

- Use Excel filters on ASIN sheet to find specific products
- Create pivot tables for custom analysis
- Export specific sheets to share with stakeholders
- Compare multiple dashboard versions to track improvements

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""

instructions_file = OUTPUT_DIR / 'Excel_Dashboard_Instructions.txt'
with open(instructions_file, 'w') as f:
    f.write(instructions)

print(f"  ✓ Saved: {instructions_file.name}")

print()
print("=" * 80)
print("✓ EXCEL DASHBOARD GENERATION COMPLETE")
print("=" * 80)
print()
print(f"Dashboard file: {excel_file}")
print(f"Instructions: {instructions_file}")
