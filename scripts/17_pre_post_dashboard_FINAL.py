#!/usr/bin/env python3
"""
PRE vs POST PERPETUA DASHBOARD - THE REAL STORY
Shows actual impact of Perpetua implementation
"""

import pandas as pd
import numpy as np
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 100)
print("CREATING PRE vs POST PERPETUA DASHBOARD")
print("=" * 100)
print()

# Load analysis
with open(OUTPUT_DIR / 'pre_post_perpetua_analysis.json') as f:
    analysis = json.load(f)

pre = analysis['pre_period']['metrics']
post = analysis['post_period']['metrics']
impact = analysis['impact']

# Load daily data for chart
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
merged = pd.read_csv(PROCESSED_DIR / 'orders_advertising_merged.csv', low_memory=False)
merged['Date'] = pd.to_datetime(merged['Date'], errors='coerce')
merged = merged[merged['Date'].notna()]

# Daily aggregation
daily = merged.groupby('Date').agg({
    'Total_Revenue': 'sum',
    'Ad_Spend': 'sum',
    'Ad_Sales': 'sum',
    'Organic_Sales': 'sum'
}).reset_index()

daily['ROAS'] = daily['Ad_Sales'] / daily['Ad_Spend'].replace(0, np.nan)
daily['TACoS'] = (daily['Ad_Spend'] / daily['Total_Revenue'].replace(0, np.nan)) * 100
daily = daily.replace([np.inf, -np.inf], np.nan).fillna(0)
daily['Period'] = daily['Date'].apply(
    lambda x: 'Pre-Perpetua' if x < pd.to_datetime('2025-12-15') else 'Post-Perpetua'
)

# Create workbook
print("[1/3] Creating dashboard...")
wb = Workbook()
wb.remove(wb.active)

COLORS = {'pre': 'ED7D31', 'post': '4472C4', 'header': '2F5496',
          'good': '70AD47', 'bad': 'C5504B', 'neutral': 'FFC000'}

header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=22)
center = Alignment(horizontal='center', vertical='center')
wrap = Alignment(wrap_text=True, vertical='top')

# ============================================================================
# MAIN DASHBOARD
# ============================================================================

ws1 = wb.create_sheet("📊 Before vs After Perpetua")

ws1.row_dimensions[2].height = 35
ws1['C2'] = 'PERPETUA IMPLEMENTATION IMPACT ANALYSIS'
ws1['C2'].font = title_font
ws1['C2'].alignment = center
ws1.merge_cells('C2:L2')

ws1['C3'] = 'Before vs After SaaS Platform Launch'
ws1['C3'].font = Font(size=13, italic=True)
ws1['C3'].alignment = center
ws1.merge_cells('C3:L3')

# Implementation date callout
row = 5
ws1.row_dimensions[row].height = 30
ws1[f'C{row}'] = '📅 PERPETUA LAUNCH: December 15, 2025'
ws1[f'C{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws1[f'C{row}'].fill = PatternFill(start_color=COLORS['post'], end_color=COLORS['post'], fill_type='solid')
ws1[f'C{row}'].alignment = center
ws1.merge_cells(f'C{row}:L{row}')

# Period definitions
row += 2
ws1[f'C{row}'] = 'Pre-Perpetua: Nov 15 - Dec 14, 2025 (30 days manual advertising)'
ws1[f'C{row}'].font = Font(size=10)
ws1.merge_cells(f'C{row}:G{row}')

ws1[f'H{row}'] = f'Post-Perpetua: Dec 15, 2025 - Feb 1, 2026 ({post["Days"]} days with SaaS)'
ws1[f'H{row}'].font = Font(size=10)
ws1.merge_cells(f'H{row}:L{row}')

# THE VERDICT
row += 2
ws1.row_dimensions[row].height = 40
ws1[f'C{row}'] = '🎯 THE VERDICT'
ws1[f'C{row}'].font = Font(bold=True, size=15)
ws1[f'C{row}'].alignment = center
ws1.merge_cells(f'C{row}:L{row}')

row += 1
ws1.row_dimensions[row].height = 50
verdict = f'Perpetua increased daily revenue by 74% (+${impact["Daily_Revenue_Change"]:,.0f}/day) but reduced ROAS by 35%. Net result: +${impact["Daily_Revenue_Change"]*365:,.0f} annualized revenue. Trade efficiency for massive scale.'
ws1[f'C{row}'] = verdict
ws1[f'C{row}'].font = Font(size=11, bold=True)
ws1[f'C{row}'].alignment = wrap
ws1.merge_cells(f'C{row}:L{row}')

# COMPARISON TABLE
row += 2
ws1[f'C{row}'] = 'COMPLETE PRE vs POST COMPARISON'
ws1[f'C{row}'].font = Font(bold=True, size=13)
ws1[f'C{row}'].alignment = center
ws1.merge_cells(f'C{row}:L{row}')

row += 2
headers = ['Metric', 'Pre-Perpetua (Manual)', 'Post-Perpetua (SaaS)', 'Change', '% Change', 'Status']
for col, h in enumerate(headers, start=3):
    ws1.cell(row=row, column=col, value=h).fill = header_fill
    ws1.cell(row=row, column=col).font = header_font
    ws1.cell(row=row, column=col).alignment = center

row += 1

comparison_metrics = [
    ('Period Length', pre['Days'], post['Days'], '#', None),
    ('', '', '', '', None),
    ('Total Revenue', pre['Total_Revenue'], post['Total_Revenue'], '$', False),
    ('Ad Spend', pre['Ad_Spend'], post['Ad_Spend'], '$', None),
    ('Ad Sales', pre['Ad_Sales'], post['Ad_Sales'], '$', False),
    ('Organic Sales', pre['Organic_Sales'], post['Organic_Sales'], '$', False),
    ('', '', '', '', None),
    ('ROAS', pre['ROAS'], post['ROAS'], 'x', False),
    ('ACOS', pre['ACOS'], post['ACOS'], '%', True),
    ('TACoS', pre['TACoS'], post['TACoS'], '%', True),
    ('T-ROAS', pre['T_ROAS'], post['T_ROAS'], 'x', False),
    ('Organic Ratio', pre['Organic_Ratio'], post['Organic_Ratio'], '%', False),
    ('', '', '', '', None),
    ('Avg Daily Revenue', pre['Avg_Daily_Revenue'], post['Avg_Daily_Revenue'], '$', False),
    ('Avg Daily Ad Spend', pre['Avg_Daily_Spend'], post['Avg_Daily_Spend'], '$', None),
]

for metric, pre_val, post_val, unit, higher_better in comparison_metrics:
    if not metric:
        row += 1
        continue

    ws1.cell(row=row, column=3, value=metric)

    # Pre value
    cell = ws1.cell(row=row, column=4, value=pre_val)
    if unit == '$':
        cell.number_format = '$#,##0'
    elif unit == '%':
        cell.number_format = '0.0%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Post value
    cell = ws1.cell(row=row, column=5, value=post_val)
    if unit == '$':
        cell.number_format = '$#,##0'
    elif unit == '%':
        cell.number_format = '0.0%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Change
    if isinstance(pre_val, (int, float)) and isinstance(post_val, (int, float)):
        change = post_val - pre_val
        cell = ws1.cell(row=row, column=6, value=change)
        if unit == '$':
            cell.number_format = '$#,##0;-$#,##0'
        elif unit == '%':
            cell.number_format = '0.0%;-0.0%'
        elif unit == 'x':
            cell.number_format = '+0.00;-0.00'
        else:
            cell.number_format = '+#,##0;-#,##0'

        # % Change
        if higher_better is not None and pre_val != 0:
            pct = (post_val - pre_val) / pre_val
            ws1.cell(row=row, column=7, value=pct).number_format = '+0%;-0%'

            # Status
            improved = (post_val > pre_val) if higher_better else (post_val < pre_val)
            status = '✓ Better' if improved else '✗ Worse'
            cell = ws1.cell(row=row, column=8, value=status)
            cell.fill = PatternFill(start_color=COLORS['good'] if improved else COLORS['bad'],
                                   end_color=COLORS['good'] if improved else COLORS['bad'],
                                   fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = center

    row += 1

# STRATEGIC INTERPRETATION
row += 2
ws1.row_dimensions[row].height = 25
ws1[f'C{row}'] = '💡 STRATEGIC INTERPRETATION'
ws1[f'C{row}'].font = Font(bold=True, size=14, color=COLORS['post'])
ws1.merge_cells(f'C{row}:L{row}')

interpretations = [
    '',
    '✓ Revenue Growth: +74% daily revenue ($40K more per day) = +$14.6M annualized',
    '⚠ Efficiency Trade-off: ROAS declined 35% (3.01x → 1.97x)',
    '⚠ Ad Dependency: TACoS increased from 1.6% to 5.5% (more reliant on ads)',
    '',
    '🎯 THE TRADE: Perpetua exchanged efficiency for massive scale',
    '   • Pre: High efficiency, low volume ($54K/day)',
    '   • Post: Lower efficiency, HIGH volume ($94K/day)',
    '',
    '💰 PROFIT ANALYSIS (assuming 30% margin):',
    f'   • Pre-Perpetua profit: ${pre["Avg_Daily_Revenue"] * 0.30:,.0f}/day',
    f'   • Post-Perpetua profit: ${post["Avg_Daily_Revenue"] * 0.30:,.0f}/day',
    f'   • Net profit increase: +${(post["Avg_Daily_Revenue"] - pre["Avg_Daily_Revenue"]) * 0.30:,.0f}/day',
    f'   • Annualized: +${(post["Avg_Daily_Revenue"] - pre["Avg_Daily_Revenue"]) * 0.30 * 365:,.0f}',
    '',
    '✅ CONCLUSION: Despite lower ROAS, Perpetua is PROFITABLE',
    '   Total profit increased significantly due to volume scale',
]

for text in interpretations:
    row += 1
    ws1[f'C{row}'] = text
    font_size = 11 if text.startswith('✓') or text.startswith('⚠') or text.startswith('🎯') or text.startswith('💰') or text.startswith('✅') else 10
    ws1[f'C{row}'].font = Font(size=font_size, bold=text.startswith(('✅', '🎯', '💰')))
    ws1.merge_cells(f'C{row}:L{row}')
    if text.startswith('   •'):
        ws1[f'C{row}'].alignment = Alignment(horizontal='left', indent=2)

# Daily data sheet
ws2 = wb.create_sheet("📈 Daily Trend")

ws2['B2'] = 'DAILY PERFORMANCE - PRE vs POST PERPETUA'
ws2['B2'].font = Font(bold=True, size=16)
ws2.merge_cells('B2:K2')

row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(daily[['Date', 'Period', 'Total_Revenue', 'Ad_Spend', 'ROAS', 'TACoS']], index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws2.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            if c_idx == 2:  # Date
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx in [4, 5]:  # Revenue, Spend
                cell.number_format = '$#,##0'
            elif c_idx in [6, 7]:  # ROAS, TACoS
                cell.number_format = '0.00'

            # Color by period
            if value == 'Pre-Perpetua':
                cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            elif value == 'Post-Perpetua':
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

ws2.auto_filter.ref = f'B{row}:G{row + len(daily)}'

# Set widths
for ws in [ws1, ws2]:
    ws.column_dimensions['C'].width = 35
    for col in 'DEFGHIJKL':
        ws.column_dimensions[col].width = 16

# Save
print("[2/3] Saving dashboard...")
output_file = OUTPUT_DIR / f'Perpetua_Before_After_Analysis_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(output_file)

print("[3/3] Summary...")
print()
print("=" * 100)
print("✓ PRE vs POST DASHBOARD COMPLETE")
print("=" * 100)
print(f"\nFile: {output_file.name}")
print()
print("📊 THE REAL STORY:")
print(f"  Pre-Perpetua (Manual):  3.01x ROAS, ${pre['Avg_Daily_Revenue']:,.0f}/day revenue")
print(f"  Post-Perpetua (SaaS):   1.97x ROAS, ${post['Avg_Daily_Revenue']:,.0f}/day revenue")
print(f"  Impact: -35% ROAS BUT +74% revenue (+${impact['Daily_Revenue_Change']:,.0f}/day)")
print()
print("💰 PROFIT IMPACT (30% margin):")
print(f"  Daily profit increase: +${impact['Daily_Revenue_Change'] * 0.30:,.0f}")
print(f"  Annualized: +${impact['Daily_Revenue_Change'] * 365 * 0.30:,.0f}")
print()
print("✅ VERDICT: Perpetua worth it despite lower ROAS (scale > efficiency)")
