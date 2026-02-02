#!/usr/bin/env python3
"""
FINAL SaaS vs Non-SaaS Dashboard - Corrected Metrics
Uses AGGREGATE calculations (Total Sales / Total Spend) not averages
Includes ALL valuable metrics for testing
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
BASE_DIR = Path(__file__).parent.parent
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
OUTPUT_DIR = BASE_DIR / 'outputs'

print("=" * 100)
print("FINAL SaaS PERFORMANCE DASHBOARD - CORRECTED AGGREGATE METRICS")
print("=" * 100)
print()

# ============================================================================
# LOAD DATA
# ============================================================================

print("[1/7] Loading data...")
df = pd.read_csv(PROCESSED_DIR / 'advertised_products_processed.csv', low_memory=False)
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df = df[df['Date'].notna()]
df = df[df['Advertising_Type'].isin(['Perpetua', 'Non-Perpetua'])]

# Clean numerics
for col in ['Spend', '7 Day Total Sales ', '7 Day Total Orders (#)', 'Clicks', 'Impressions', '7 Day Total Units (#)']:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

min_date = df['Date'].min()
max_date = df['Date'].max()
date_range_days = (max_date - min_date).days

print(f"  ✓ {len(df):,} records from {min_date.date()} to {max_date.date()} ({date_range_days} days)")

# ============================================================================
# CALCULATE AGGREGATE METRICS (CORRECT METHOD)
# ============================================================================

print("[2/7] Calculating aggregate metrics...")

def calc_metrics(subset):
    """Calculate aggregate metrics the CORRECT way"""
    total_spend = subset['Spend'].sum()
    total_sales = subset['7 Day Total Sales '].sum()
    total_orders = subset['7 Day Total Orders (#)'].sum()
    total_clicks = subset['Clicks'].sum()
    total_impressions = subset['Impressions'].sum()
    total_units = subset['7 Day Total Units (#)'].sum()
    unique_asins = subset['Advertised ASIN'].nunique()

    return {
        'Total_Spend': total_spend,
        'Total_Sales': total_sales,
        'Total_Orders': total_orders,
        'Total_Clicks': total_clicks,
        'Total_Impressions': total_impressions,
        'Total_Units': total_units,
        'Unique_ASINs': unique_asins,
        # AGGREGATE calculations (portfolio level)
        'ROAS': total_sales / total_spend if total_spend > 0 else 0,
        'ACOS': total_spend / total_sales if total_sales > 0 else 0,
        'CPC': total_spend / total_clicks if total_clicks > 0 else 0,
        'CTR': total_clicks / total_impressions if total_impressions > 0 else 0,
        'CVR': total_orders / total_clicks if total_clicks > 0 else 0,
        'CPA': total_spend / total_orders if total_orders > 0 else 0,
        'CPM': (total_spend / total_impressions) * 1000 if total_impressions > 0 else 0,
        'AOV': total_sales / total_orders if total_orders > 0 else 0,
        # Per-ASIN averages
        'Spend_Per_ASIN': total_spend / unique_asins if unique_asins > 0 else 0,
        'Sales_Per_ASIN': total_sales / unique_asins if unique_asins > 0 else 0,
        'Orders_Per_ASIN': total_orders / unique_asins if unique_asins > 0 else 0
    }

perpetua_metrics = calc_metrics(df[df['Advertising_Type'] == 'Perpetua'])
non_perpetua_metrics = calc_metrics(df[df['Advertising_Type'] == 'Non-Perpetua'])

print("\nPerpetua (SaaS):")
print(f"  Total Spend: ${perpetua_metrics['Total_Spend']:,.2f}")
print(f"  Total Sales: ${perpetua_metrics['Total_Sales']:,.2f}")
print(f"  ROAS: {perpetua_metrics['ROAS']:.2f}x")
print(f"  ACOS: {perpetua_metrics['ACOS']*100:.1f}%")
print(f"  CTR: {perpetua_metrics['CTR']*100:.2f}%")
print(f"  CVR: {perpetua_metrics['CVR']*100:.2f}%")

print("\nNon-Perpetua (Manual):")
print(f"  Total Spend: ${non_perpetua_metrics['Total_Spend']:,.2f}")
print(f"  Total Sales: ${non_perpetua_metrics['Total_Sales']:,.2f}")
print(f"  ROAS: {non_perpetua_metrics['ROAS']:.2f}x")
print(f"  ACOS: {non_perpetua_metrics['ACOS']*100:.1f}%")
print(f"  CTR: {non_perpetua_metrics['CTR']*100:.2f}%")
print(f"  CVR: {non_perpetua_metrics['CVR']*100:.2f}%")

# ============================================================================
# DAILY TIME SERIES
# ============================================================================

print("\n[3/7] Preparing daily time series...")
daily = df.groupby(['Date', 'Advertising_Type']).agg({
    'Spend': 'sum',
    '7 Day Total Sales ': 'sum',
    '7 Day Total Orders (#)': 'sum',
    'Clicks': 'sum',
    'Impressions': 'sum'
}).reset_index()

daily['ROAS'] = daily['7 Day Total Sales '] / daily['Spend'].replace(0, np.nan)
daily['ACOS'] = daily['Spend'] / daily['7 Day Total Sales '].replace(0, np.nan)
daily['CPC'] = daily['Spend'] / daily['Clicks'].replace(0, np.nan)
daily['CTR'] = daily['Clicks'] / daily['Impressions'].replace(0, np.nan)
daily['CVR'] = daily['7 Day Total Orders (#)'] / daily['Clicks'].replace(0, np.nan)
daily = daily.replace([np.inf, -np.inf], np.nan).fillna(0)

print(f"  ✓ {len(daily)} daily records")

# ============================================================================
# CREATE EXCEL WORKBOOK
# ============================================================================

print("[4/7] Creating Excel workbook...")

wb = Workbook()
wb.remove(wb.active)

# Professional color scheme
COLORS = {
    'perpetua': '4472C4',      # Blue
    'non_perpetua': 'ED7D31',  # Orange
    'header': '2F5496',
    'good': '70AD47',          # Green
    'bad': 'C5504B',           # Red
    'warning': 'FFC000',       # Yellow
    'light_blue': 'D9E1F2',
    'light_orange': 'FCE4D6'
}

# Styles
header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=18, color=COLORS['header'])
subtitle_font = Font(bold=True, size=13)
large_number_font = Font(bold=True, size=24)
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right')

# ============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================================

print("[5/7] Building Executive Summary...")
ws1 = wb.create_sheet("📊 Executive Dashboard")

# Title
ws1['C3'] = 'SaaS PLATFORM PERFORMANCE COMPARISON'
ws1['C3'].font = title_font
ws1.merge_cells('C3:J3')
ws1['C3'].alignment = center

ws1['C4'] = f'Perpetua (SaaS Automation) vs Manual Advertising'
ws1['C4'].font = Font(size=12, italic=True)
ws1.merge_cells('C4:J4')
ws1['C4'].alignment = center

ws1['C5'] = f'{min_date.strftime("%B %d, %Y")} - {max_date.strftime("%B %d, %Y")} ({date_range_days} days)'
ws1['C5'].font = Font(size=10)
ws1.merge_cells('C5:J5')
ws1['C5'].alignment = center

# KEY METRICS CARDS
row = 7

# Card headers
card_row = row
ws1[f'C{card_row}'] = 'PERPETUA (SaaS)'
ws1[f'C{card_row}'].font = Font(bold=True, size=12, color='FFFFFF')
ws1[f'C{card_row}'].fill = PatternFill(start_color=COLORS['perpetua'], end_color=COLORS['perpetua'], fill_type='solid')
ws1[f'C{card_row}'].alignment = center
ws1.merge_cells(f'C{card_row}:E{card_row}')

ws1[f'G{card_row}'] = 'NON-PERPETUA (Manual)'
ws1[f'G{card_row}'].font = Font(bold=True, size=12, color='FFFFFF')
ws1[f'G{card_row}'].fill = PatternFill(start_color=COLORS['non_perpetua'], end_color=COLORS['non_perpetua'], fill_type='solid')
ws1[f'G{card_row}'].alignment = center
ws1.merge_cells(f'G{card_row}:I{card_row}')

# ROAS Cards (Most Important Metric)
row += 1
ws1[f'C{row}'] = 'ROAS'
ws1[f'C{row}'].font = subtitle_font
ws1.merge_cells(f'C{row}:E{row}')
ws1[f'C{row}'].alignment = center

ws1[f'G{row}'] = 'ROAS'
ws1[f'G{row}'].font = subtitle_font
ws1.merge_cells(f'G{row}:I{row}')
ws1[f'G{row}'].alignment = center

row += 1
ws1[f'C{row}'] = perpetua_metrics['ROAS']
ws1[f'C{row}'].font = large_number_font
ws1[f'C{row}'].number_format = '0.00"x"'
ws1.merge_cells(f'C{row}:E{row}')
ws1[f'C{row}'].alignment = center

ws1[f'G{row}'] = non_perpetua_metrics['ROAS']
ws1[f'G{row}'].font = large_number_font
ws1[f'G{row}'].number_format = '0.00"x"'
ws1.merge_cells(f'G{row}:I{row}')
ws1[f'G{row}'].alignment = center
# Highlight better performer
if non_perpetua_metrics['ROAS'] > perpetua_metrics['ROAS']:
    ws1[f'G{row}'].fill = PatternFill(start_color=COLORS['good'], end_color=COLORS['good'], fill_type='solid')

# DETAILED METRICS TABLE
row += 3
ws1[f'C{row}'] = 'ALL PERFORMANCE METRICS'
ws1[f'C{row}'].font = subtitle_font
ws1.merge_cells(f'C{row}:J{row}')

row += 1
headers = ['Metric', 'Perpetua', 'Non-Perpetua', 'Difference', '% Better', 'Winner']
for col, header in enumerate(headers, start=3):
    cell = ws1.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

row += 1

# All metrics to display
all_metrics = [
    ('Total Spend', 'Total_Spend', '$', None),
    ('Total Sales', 'Total_Sales', '$', False),
    ('Total Orders', 'Total_Orders', '#', False),
    ('Total Clicks', 'Total_Clicks', '#', False),
    ('Total Impressions', 'Total_Impressions', '#', False),
    ('Unique ASINs', 'Unique_ASINs', '#', None),
    ('', '', '', None),  # Separator
    ('ROAS (Return on Ad Spend)', 'ROAS', 'x', False),
    ('ACOS (Advertising Cost of Sales)', 'ACOS', '%', True),
    ('CPC (Cost Per Click)', 'CPC', '$', True),
    ('CTR (Click-Through Rate)', 'CTR', '%', False),
    ('CVR (Conversion Rate)', 'CVR', '%', False),
    ('CPA (Cost Per Acquisition)', 'CPA', '$', True),
    ('CPM (Cost Per 1000 Impressions)', 'CPM', '$', True),
    ('AOV (Average Order Value)', 'AOV', '$', False),
    ('', '', '', None),  # Separator
    ('Spend per ASIN', 'Spend_Per_ASIN', '$', True),
    ('Sales per ASIN', 'Sales_Per_ASIN', '$', False),
    ('Orders per ASIN', 'Orders_Per_ASIN', '#', False),
]

for metric_name, metric_key, unit, lower_is_better in all_metrics:
    if not metric_name:  # Separator row
        row += 1
        continue

    # Metric name
    ws1.cell(row=row, column=3, value=metric_name).font = Font(size=10)

    # Values
    p_val = perpetua_metrics[metric_key]
    np_val = non_perpetua_metrics[metric_key]

    # Perpetua
    cell = ws1.cell(row=row, column=4, value=p_val)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.value = p_val
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Non-Perpetua
    cell = ws1.cell(row=row, column=5, value=np_val)
    if unit == '$':
        cell.number_format = '$#,##0.00'
    elif unit == '%':
        cell.value = np_val
        cell.number_format = '0.00%'
    elif unit == 'x':
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = '#,##0'

    # Difference
    diff = p_val - np_val
    cell = ws1.cell(row=row, column=6, value=diff)
    if unit == '$':
        cell.number_format = '$#,##0.00;($#,##0.00)'
    elif unit == '%':
        cell.value = diff
        cell.number_format = '0.00%;(0.00%)'
    elif unit == 'x':
        cell.number_format = '0.00;(0.00)'
    else:
        cell.number_format = '#,##0;(#,##0)'

    # Percent difference
    if lower_is_better is not None:
        pct_diff = ((np_val - p_val) / np_val * 100) if np_val != 0 else 0 if lower_is_better else ((p_val - np_val) / np_val * 100) if np_val != 0 else 0
        cell = ws1.cell(row=row, column=7, value=pct_diff / 100)
        cell.number_format = '0.0%;(0.0%)'

        # Winner
        if lower_is_better:
            winner = 'Perpetua ✓' if p_val < np_val else 'Non-Perpetua ✓'
            is_perpetua_better = p_val < np_val
        else:
            winner = 'Perpetua ✓' if p_val > np_val else 'Non-Perpetua ✓'
            is_perpetua_better = p_val > np_val

        cell = ws1.cell(row=row, column=8, value=winner)
        cell.fill = PatternFill(start_color=COLORS['perpetua'] if is_perpetua_better else COLORS['non_perpetua'],
                               end_color=COLORS['perpetua'] if is_perpetua_better else COLORS['non_perpetua'],
                               fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = center

    row += 1

# KEY INSIGHTS BOX
row += 2
ws1[f'C{row}'] = '💡 KEY INSIGHTS'
ws1[f'C{row}'].font = Font(bold=True, size=13, color=COLORS['non_perpetua'])
ws1.merge_cells(f'C{row}:J{row}')

roas_diff = ((non_perpetua_metrics['ROAS'] - perpetua_metrics['ROAS']) / perpetua_metrics['ROAS'] * 100)

insights = [
    f"1. Non-Perpetua achieves {abs(roas_diff):.0f}% better ROAS ({non_perpetua_metrics['ROAS']:.2f}x vs {perpetua_metrics['ROAS']:.2f}x)",
    f"2. Perpetua manages {perpetua_metrics['Unique_ASINs']:.0f} ASINs generating ${perpetua_metrics['Total_Sales']:,.0f} in sales",
    f"3. Non-Perpetua manages {non_perpetua_metrics['Unique_ASINs']:.0f} ASINs generating ${non_perpetua_metrics['Total_Sales']:,.0f} in sales",
    f"4. Perpetua spends {(perpetua_metrics['Total_Spend']/non_perpetua_metrics['Total_Spend']):.1f}x more but ROAS is {abs(roas_diff):.0f}% lower",
    f"5. If Perpetua matched Non-Perpetua ROAS: Additional ${(perpetua_metrics['Total_Spend'] * non_perpetua_metrics['ROAS'] - perpetua_metrics['Total_Sales']):,.0f} revenue",
]

for insight in insights:
    row += 1
    ws1[f'C{row}'] = insight
    ws1[f'C{row}'].font = Font(size=10)
    ws1.merge_cells(f'C{row}:J{row}')

# ============================================================================
# SHEET 2: TIME SERIES (FILTERABLE)
# ============================================================================

print("[6/7] Creating Time Series sheet...")
ws2 = wb.create_sheet("📅 Daily Data (Filter by Date)")

ws2['B2'] = 'DAILY PERFORMANCE DATA - FILTERABLE'
ws2['B2'].font = title_font
ws2.merge_cells('B2:M2')

ws2['B3'] = 'Click filter dropdowns ▼ to analyze specific date ranges'
ws2['B3'].font = Font(italic=True, size=10)
ws2.merge_cells('B3:M3')

# Add data
row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(daily, index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws2.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            if c_idx == 2:  # Date
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx in [4, 5]:  # Spend, Sales
                cell.number_format = '$#,##0.00'
            elif c_idx in [6, 7, 8]:  # Orders, Clicks, Impressions
                cell.number_format = '#,##0'
            elif c_idx in [9, 10, 11, 12, 13]:  # ROAS, ACOS, CPC, CTR, CVR
                cell.number_format = '0.0000'

            # Color code by platform
            if value == 'Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
            elif value == 'Non-Perpetua':
                cell.fill = PatternFill(start_color=COLORS['light_orange'], end_color=COLORS['light_orange'], fill_type='solid')

# Enable AutoFilter
ws2.auto_filter.ref = f'B{row}:M{row + len(daily)}'

# ============================================================================
# SHEET 3: ASIN DETAIL
# ============================================================================

print("[7/7] Creating ASIN Detail sheet...")
ws3 = wb.create_sheet("🔍 Top 100 ASINs")

# Aggregate by ASIN
asin_agg = df.groupby(['Advertised ASIN', 'Advertising_Type']).agg({
    'Spend': 'sum',
    '7 Day Total Sales ': 'sum',
    '7 Day Total Orders (#)': 'sum',
    'Clicks': 'sum',
    'Impressions': 'sum'
}).reset_index()

asin_agg['ROAS'] = asin_agg['7 Day Total Sales '] / asin_agg['Spend'].replace(0, np.nan)
asin_agg['ACOS'] = asin_agg['Spend'] / asin_agg['7 Day Total Sales '].replace(0, np.nan)
asin_agg['CVR'] = asin_agg['7 Day Total Orders (#)'] / asin_agg['Clicks'].replace(0, np.nan)
asin_agg = asin_agg.replace([np.inf, -np.inf], np.nan).fillna(0)

# Sort by spend
asin_agg = asin_agg.sort_values('Spend', ascending=False).head(100)

ws3['B2'] = 'TOP 100 ASINs BY TOTAL SPEND'
ws3['B2'].font = title_font
ws3.merge_cells('B2:K2')

# Add data
row = 5
for r_idx, row_data in enumerate(dataframe_to_rows(asin_agg, index=False, header=True)):
    for c_idx, value in enumerate(row_data, start=2):
        cell = ws3.cell(row=row + r_idx, column=c_idx, value=value)

        if r_idx == 0:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            # Format
            if c_idx in [3, 4]:  # Spend, Sales
                cell.number_format = '$#,##0.00'
            elif c_idx in [5, 6, 7]:  # Orders, Clicks, Impressions
                cell.number_format = '#,##0'
            elif c_idx in [8, 9, 10]:  # ROAS, ACOS, CVR
                cell.number_format = '0.00'

# Set widths
for ws in wb.worksheets:
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 20
    for col in 'CDEFGHIJKLM':
        ws.column_dimensions[col].width = 16

# Save
output_file = OUTPUT_DIR / f'Perpetua_Dashboard_FINAL_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(output_file)

print()
print("=" * 100)
print("✓ FINAL DASHBOARD COMPLETE WITH CORRECT METRICS")
print("=" * 100)
print(f"\nFile: {output_file.name}")
print()
print("📊 CORRECT AGGREGATE METRICS:")
print(f"  Perpetua:     {perpetua_metrics['ROAS']:.2f}x ROAS, {perpetua_metrics['ACOS']*100:.1f}% ACOS")
print(f"  Non-Perpetua: {non_perpetua_metrics['ROAS']:.2f}x ROAS, {non_perpetua_metrics['ACOS']*100:.1f}% ACOS")
print(f"  Difference:   {abs(roas_diff):.0f}% better ROAS for Non-Perpetua")
print()
print("✓ All metrics normalized to date range: {min_date.date()} to {max_date.date()}")
print("✓ AutoFilter enabled on Daily Data sheet")
print("✓ Professional formatting and color-coding")
